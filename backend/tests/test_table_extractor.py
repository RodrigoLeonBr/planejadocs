"""Testes do extrator de tabelas (spec convert.md)."""
import json

from openpyxl import load_workbook

from app.core.table_extractor import (
    extract_tables,
    save_excel,
    tables_to_csv,
    tables_to_json,
)


def test_extrai_tabela(table_pdf):
    tables = extract_tables(str(table_pdf))
    assert len(tables) == 1
    t = tables[0]
    assert t["page"] == 1
    assert t["table_index"] == 0
    assert t["rows"] == [["Nome", "Valor"], ["UBS A", "100"], ["UBS B", "200"]]


def test_sem_tabela_retorna_lista_vazia(native_pdf):
    assert extract_tables(str(native_pdf)) == []


def test_tables_to_json(table_pdf):
    tables = extract_tables(str(table_pdf))
    parsed = json.loads(tables_to_json(tables))
    assert parsed == tables


def test_tables_to_csv(table_pdf):
    tables = extract_tables(str(table_pdf))
    csv_text = tables_to_csv(tables)
    assert "Nome,Valor" in csv_text
    assert "UBS A,100" in csv_text


def test_save_excel(table_pdf, tmp_path):
    tables = extract_tables(str(table_pdf))
    out = tmp_path / "tabelas.xlsx"
    save_excel(tables, str(out))
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value == "Nome"
    assert ws["B3"].value == "200"
